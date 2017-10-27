#!/usr/bin/python
# -*- coding: utf-8 -*-	

import os.path
import json

from openpyxl import Workbook
from openpyxl import load_workbook

DIR_TABLES = os.path.join("tables/")
DIR_JSON = os.path.join("json/")

# READ ALL ROWS IN WORKSHEET AND TRANSFORM INTO JSON
def all_data_to_json(worksheet, filename, sheetname):
	with open('{}{}_{}.json'.format(DIR_JSON,filename,sheetname), 'w') as file:
		json_data = []

		for row in range(1, worksheet.max_row):
			item = {}
			for column in range(worksheet.max_column):
				try:
					item[worksheet.cell(row=1, column=column+1).value.upper()] = worksheet.cell(row=row+1, column=column+1).value.encode('utf-8')
				except:
					try:
						item[worksheet.cell(row=1, column=column+1).value.upper()] = worksheet.cell(row=row+1, column=column+1).value
					except:
						item[worksheet.cell(row=1, column=column+1).value] = worksheet.cell(row=row+1, column=column+1).value
			json_data.append(item)

		json.dump(json_data, file, indent = 4, ensure_ascii = False) # sort_keys = True
		file.close()


# READ ALL FILES INSIDE TABLES FOLDER
def read_tables_folder():
	files = []
	directory = os.listdir(DIR_TABLES)
	
	for file in directory:
		if(not ".json" in file):
			files.append(file)
	
	return files

# CREATE TABLES FOLDER AND JSON FOLDER
def create_dependencies():
	if(not os.path.exists(DIR_TABLES)):
		os.makedirs(DIR_TABLES)
	
	if(not os.path.exists(DIR_JSON)):
		os.makedirs(DIR_JSON)	

def main():
	
	create_dependencies()
	files = read_tables_folder();

	if(not files):
		print "no excel file inside tables folder"

	for file in files:
		try:
			wb = load_workbook(filename=DIR_TABLES + file , read_only=True)
		except:
			print(file + " not supported.")
			return
		try:
			filename = file.split('.')

			sheets = wb.get_sheet_names()
			for sheet in sheets:
				ws = wb[sheet]

				all_data_to_json(ws, filename[0], sheet)
		except:
			print("script error")

if __name__ == '__main__':
	main()